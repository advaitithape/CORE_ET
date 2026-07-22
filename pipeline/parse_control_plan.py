"""Parse the Control Plan .xlsx -> per-operation control lines (one sheet per op).

Column layout (data rows after the ~row-20 header):
  c1  Part Process Number (op no, merged per sheet)
  c2  Process Name / Operation
  c3  Machine / Device / Tools
  c4  Sr. No.
  c5  Product characteristic
  c6  Process characteristic
  c7  Special Character / Class
  c8  Product/Process Specification / Tolerance
  c9  Measurement / Evaluation method (gauge)
  c10 Sample size
  c11 Frequency
  c12 Responsibility
  c13 Record
  c14 Control Method
  c15 Error Proofing
  c16 Reaction Plan
"""
import warnings
warnings.filterwarnings("ignore")
import openpyxl
from .common import norm_op, clean, xlsx_merged_map, xlsx_cell

C_OP, C_OPNAME, C_MACHINE, C_SR = 1, 2, 3, 4
C_PROD, C_PROC, C_SPECIAL, C_SPEC = 5, 6, 7, 8
C_METHOD, C_SIZE, C_FREQ, C_RESP, C_RECORD = 9, 10, 11, 12, 13
C_CTRLMETHOD, C_ERRPROOF, C_REACT = 14, 15, 16


def _find_cp_no(ws):
    for r in range(1, min(15, ws.max_row) + 1):
        for c in range(1, min(12, ws.max_column) + 1):
            v = clean(ws.cell(r, c).value)
            if "CONTROL PLAN NO" in v.upper():
                for cc in range(c + 1, min(c + 6, ws.max_column) + 1):
                    nxt = clean(ws.cell(r, cc).value).lstrip("'")
                    if nxt and "CP/" in nxt.upper():
                        return nxt
    return None


def _header_row(ws):
    """Locate the 'Sr.No.' header row so data parsing starts right after it."""
    for r in range(1, min(30, ws.max_row) + 1):
        if clean(ws.cell(r, C_SR).value).lower().startswith("sr"):
            return r
    return 20


def parse(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    controls = []
    sheets = []
    for ws in wb.worksheets:
        cp_no = _find_cp_no(ws)
        hdr = _header_row(ws)
        mmap = xlsx_merged_map(ws)
        # op number: prefer the merged part-process-number column, else sheet name
        op_from_col = None
        for r in range(hdr + 1, ws.max_row + 1):
            op_from_col = norm_op(xlsx_cell(ws, mmap, r, C_OP))
            if op_from_col is not None:
                break
        op_no = op_from_col if op_from_col is not None else norm_op(ws.title)
        sheets.append({"sheet": ws.title, "cp_no": cp_no, "op_no": op_no})

        for r in range(hdr + 1, ws.max_row + 1):
            prod = clean(xlsx_cell(ws, mmap, r, C_PROD))
            proc = clean(xlsx_cell(ws, mmap, r, C_PROC))
            spec = clean(xlsx_cell(ws, mmap, r, C_SPEC))
            sr = clean(xlsx_cell(ws, mmap, r, C_SR))
            char = prod or proc
            if not char or char in ("----", "--", "-"):
                continue
            # skip sub-header rows like 'Chemical Composition:-'
            if not spec and not sr.strip().isdigit():
                continue
            controls.append({
                "op_no": op_no,
                "cp_no": cp_no,
                "operation_name": clean(xlsx_cell(ws, mmap, r, C_OPNAME)),
                "sr_no": sr,
                "characteristic": char,
                "char_type": "product" if prod else "process",
                "special": clean(xlsx_cell(ws, mmap, r, C_SPECIAL)),
                "specification": spec,
                "measurement_method": clean(xlsx_cell(ws, mmap, r, C_METHOD)),
                "sample_size": clean(xlsx_cell(ws, mmap, r, C_SIZE)),
                "frequency": clean(xlsx_cell(ws, mmap, r, C_FREQ)),
                "responsibility": clean(xlsx_cell(ws, mmap, r, C_RESP)),
                "reaction_plan": clean(xlsx_cell(ws, mmap, r, C_REACT)),
                "source": {"doc": "ControlPlan", "sheet": ws.title, "row": r},
            })
    return {"sheets": sheets, "controls": controls}


if __name__ == "__main__":
    import json, sys
    out = parse(sys.argv[1] if len(sys.argv) > 1 else
                "data/CE21609_CONTROL PLAN 2.xlsx")
    print(f"sheets={len(out['sheets'])} controls={len(out['controls'])}")
    for s in out["sheets"]:
        print(" ", s)
    print(json.dumps(out["controls"][:3], indent=2, ensure_ascii=False))
